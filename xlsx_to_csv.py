import openpyxl
from openpyxl import utils
import datetime
import pandas as pd
import os


def xlsx_to_csv(user_name='加納', file_name='test.xlsx'):
    """
    [矢作・ハマ]の様になっているセルが存在していてそれには対応していない
    ランチに対応していない

    :param user_name: 検索したい名前
    :param file_name: 読み込むファイル名
    :return: 出勤時間がカラムで出勤する店の名前がインデックスのcsvファイル
    """
    wb = openpyxl.load_workbook(f'{file_name}')
    sheet = wb['Sheet2']
    range_idx_boxs_itigou = [[i * 2 + 1, j * 6 + 4, i * 2 + 2, j * 6 + 9] for j in range(5) for i in
                             range(7)]  # 検索するシフトの塊のリスト(一号店)
    range_idx_boxs_hanare = [[i * 2 + 1, j * 6 + 38, i * 2 + 2, j * 6 + 43] for j in range(5) for i in
                             range(7)]  # 検索するシフトの塊のリスト(はなれ)
    for range_idx in range_idx_boxs_itigou:
        range_idx[0] = utils.cell._get_column_letter(range_idx[0])
        range_idx[2] = utils.cell._get_column_letter(range_idx[2])

    for range_idx in range_idx_boxs_hanare:
        range_idx[0] = utils.cell._get_column_letter(range_idx[0])
        range_idx[2] = utils.cell._get_column_letter(range_idx[2])

    def search_name(range_idx_boxs, user_name='岩崎'):
        """
        :param range_idx_boxs: 検索するシフトの塊のリスト
        :param user_name:検索したい名前
        :return: 入るシフトの時刻のリスト
        """
        go_work_days = []  # 検索した人が働く日にちと時間
        month = utils.datetime.from_excel(sheet['A1'].value)  # エクセルファイルの月
        for range_idx in range_idx_boxs:
            cell_range_str = range_idx[0] + str(range_idx[1]) + ':' + range_idx[2] + str(range_idx[3])  # 読み出す範囲
            box = sheet[cell_range_str]  # 取り出す日にちのシフトの塊
            day = str(box[0][0].value)  # 取り出した日にち
            lunch_member = box[1][1].value  # ランチメンバー
            if lunch_member != None:
                lunch_member = lunch_member.split('・')  # ランチメンバーをリスト一人一人リストで格納
            dinner_member_ex = [box[3][0].value, box[4][0].value, box[5][0].value]
            dinner_member = []
            for cell in dinner_member_ex:  # cellはディナーの人の名前が入ってるセル
                if cell != None and str(cell) != '16:00:00':
                    if f'{user_name}' in cell:
                        if ':' in cell:
                            time = cell.split(f'{user_name}')[1]
                            hour = time.split(':00')[0]
                            day = month + datetime.timedelta(days=int(day) - 1) + datetime.timedelta(hours=int(hour))
                        else:
                            day = month + datetime.timedelta(days=int(day) - 1) + datetime.timedelta(hours=int(16))
                        go_work_days.append(day)
        return go_work_days

    def write_csv(go_work_days, shop, user_name):
        df = pd.DataFrame(data=go_work_days, index=[f'{shop}' for i in range(len(go_work_days))],
                          columns=[f'{user_name}の出勤時間'])
        return df

    go_work_days_itigou = search_name(range_idx_boxs_itigou, user_name)
    df_itigou = write_csv(go_work_days_itigou, shop='あかり一号店', user_name=user_name)
    go_work_days_hanare = search_name(range_idx_boxs_hanare, user_name)
    df_hanare = write_csv(go_work_days_hanare, shop='あかりはなれ', user_name=user_name)
    df = pd.concat([df_itigou, df_hanare])

    csv_file_name = 'csv/' + file_name.split(".xlsx")[0] + f'_{user_name}.csv'  # 人ごとのcsvファイルのpath名を作成

    """
    csvファイルが存在している場合は一度GoogleCalendarに書き込んでいるので、二度目にならない様にファイルの存在確認
    """
    if not os.path.exists(csv_file_name):
        df.to_csv(csv_file_name)
        print("csvファイルを新規作成しました")
        return df
    else:
        print('csvファイルが存在しています')


if __name__ == '__main__':
    df = xlsx_to_csv('加納')
